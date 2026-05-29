import os
import csv
import uuid
from datetime import datetime

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from api_client import fetch_api_rows, load_api_config, save_api_config
from constants import (
    ADDRESS_FIELDS,
    BANK_CODES_SHEET,
    CLEAN_TEXT_FIELDS,
    COLUMNS,
    DATA_START_ROW,
    MANDATORY_FIELDS,
    OUTPUT_DIR,
    TEMPLATE_PATH,
    TEMPLATE_SHEET,
    UPLOAD_DIR,
)
from utils import (
    canonical_bank,
    canonical_lookup,
    clean_account,
    clean_amount,
    clean_text,
    is_blank,
    map_row_keys,
    normalize_header,
)


def load_bank_names():
    wb = load_workbook(TEMPLATE_PATH, read_only=True, keep_vba=True, data_only=True)
    ws = wb[BANK_CODES_SHEET]
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and row[1]:
            names.add(clean_text(row[1]))
    return names


def read_csv_rows(path):
    encodings = ["utf-8-sig", "cp1252", "latin1"]
    last_error = None
    for enc in encodings:
        try:
            with open(path, newline="", encoding=enc) as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error


def read_excel_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    lookup = canonical_lookup()
    best_row, best_score = 1, -1

    for row_num in range(1, min(ws.max_row, 25) + 1):
        values = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]
        score = sum(1 for value in values if normalize_header(value) in lookup)
        if score > best_score:
            best_row, best_score = row_num, score

    headers = [ws.cell(best_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []

    for r in range(best_row + 1, ws.max_row + 1):
        record, has_any = {}, False
        for c, header in enumerate(headers, start=1):
            if header is None:
                continue
            value = ws.cell(r, c).value
            if value not in (None, ""):
                has_any = True
            record[str(header)] = value
        if has_any:
            rows.append(record)

    return rows


def read_input_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return read_csv_rows(path)
    return read_excel_rows(path)


def validate_and_clean(rows):
    bank_names = load_bank_names()
    cleaned_rows, error_rows, all_unmatched = [], [], set()

    for idx, row in enumerate(rows, start=2):
        out, unmatched = map_row_keys(row)
        all_unmatched.update(unmatched)
        errors = []

        out["Amount"] = clean_amount(out["Amount"])
        out["Beneficiary Account Number"] = clean_account(out["Beneficiary Account Number"])
        out["Beneficiary Bank"] = canonical_bank(out["Beneficiary Bank"])
        out["Type of Beneficiary"] = clean_text(out["Type of Beneficiary"])
        out["Purpose Code"] = clean_text(out["Purpose Code"])
        out["RTGS Purpose Code"] = clean_text(out["RTGS Purpose Code"])

        for field in CLEAN_TEXT_FIELDS:
            max_len = 35 if field in ADDRESS_FIELDS else None
            out[field] = clean_text(out[field], max_len)

        for field in MANDATORY_FIELDS:
            if is_blank(out[field]):
                errors.append(f"Missing mandatory field: {field}")

        benef_type = out["Type of Beneficiary"]
        if benef_type.startswith("IND") and is_blank(out["Beneficiary First Name"]):
            errors.append("Missing Beneficiary First Name for Individual beneficiary")
        if benef_type.startswith("CORP"):
            out["Beneficiary First Name"] = ""
            out["Beneficiary Middle Name"] = ""

        if not is_blank(out["Amount"]):
            try:
                if float(out["Amount"]) <= 0:
                    errors.append("Amount must be greater than zero")
            except Exception:
                errors.append("Amount must be numeric")

        if not is_blank(out["Beneficiary Bank"]) and out["Beneficiary Bank"] not in bank_names:
            errors.append("Beneficiary Bank not found in Bank Codes sheet")

        cleaned_rows.append(out)
        if errors:
            error_rows.append({"Source Row": idx, "Errors": "; ".join(errors), **out})

    return cleaned_rows, error_rows, sorted(all_unmatched)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A2"


def write_clean_output(cleaned_rows, error_rows, original_filename):
    wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb[TEMPLATE_SHEET]
    clear_until = max(ws.max_row, DATA_START_ROW + len(cleaned_rows) + 10)

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=clear_until, min_col=1, max_col=20):
        for cell in row:
            cell.value = None

    for r_idx, item in enumerate(cleaned_rows, start=DATA_START_ROW):
        for c_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(r_idx, c_idx).value = item.get(col, "")

    for sheet_name in ["Validation Errors", "Summary"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    err_ws = wb.create_sheet("Validation Errors")
    err_headers = ["Source Row", "Errors"] + COLUMNS
    err_ws.append(err_headers)
    for err in error_rows:
        err_ws.append([err.get(h, "") for h in err_headers])
    style_sheet(err_ws)

    sum_ws = wb.create_sheet("Summary", 0)
    sum_ws.append(["Metric", "Value"])
    sum_ws.append(["Source File", original_filename])
    sum_ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    sum_ws.append(["Total Rows Processed", len(cleaned_rows)])
    sum_ws.append(["Rows With Errors", len(error_rows)])
    sum_ws.append(["Ready Rows", len(cleaned_rows) - len(error_rows)])
    style_sheet(sum_ws)

    output_path = os.path.join(OUTPUT_DIR, f"bdo_cleaned_upload_{uuid.uuid4().hex[:8]}.xlsm")
    wb.save(output_path)
    return output_path


def show_results(cleaned_rows, error_rows, unmatched, output_path):
    st.success("Processing complete.")

    col1, col2, col3 = st.columns(3)
    col1.metric("Total Rows", len(cleaned_rows))
    col2.metric("Rows With Errors", len(error_rows))
    col3.metric("Ready Rows", len(cleaned_rows) - len(error_rows))

    if unmatched:
        st.warning("Unmatched input columns: " + ", ".join(unmatched))

    if error_rows:
        st.subheader("Validation Error Preview")
        st.dataframe(error_rows[:50], use_container_width=True)

    with open(output_path, "rb") as f:
        st.download_button(
            "Download Cleaned BDO Upload File",
            data=f,
            file_name=os.path.basename(output_path),
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
        )


def upload_tab():
    uploaded_file = st.file_uploader("Upload CSV, XLSX, or XLSM", type=["csv", "xlsx", "xlsm"])

    if uploaded_file and st.button("Validate Uploaded File", type="primary"):
        saved_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4().hex}_{uploaded_file.name}")
        with open(saved_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        try:
            rows = read_input_file(saved_path)
            cleaned_rows, error_rows, unmatched = validate_and_clean(rows)
            output_path = write_clean_output(cleaned_rows, error_rows, uploaded_file.name)
            show_results(cleaned_rows, error_rows, unmatched, output_path)
        except Exception as exc:
            st.error(f"Processing failed: {exc}")


def api_tab():
    config = load_api_config()
    st.caption("For Smartsheet, use: https://api.smartsheet.com/2.0/sheets/YOUR_SHEET_ID")

    api_url = st.text_input("API URL", value=config.get("api_url", ""))
    bearer_token = st.text_input("Bearer Token", value=config.get("bearer_token", ""), type="password")
    json_data_path = st.text_input("JSON Data Path", value=config.get("json_data_path", ""))
    custom_headers = st.text_area("Custom Headers", value=config.get("custom_headers", ""))

    current = {
        "api_url": api_url,
        "method": "GET",
        "auth_type": "bearer",
        "bearer_token": bearer_token,
        "basic_username": "",
        "basic_password": "",
        "custom_headers": custom_headers,
        "json_data_path": json_data_path,
    }

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Save API Config"):
            try:
                save_api_config(current)
                st.success("API configuration saved.")
            except Exception as exc:
                st.error(f"Saving failed: {exc}")

    with col2:
        if st.button("Fetch and Validate API Data", type="primary"):
            try:
                rows = fetch_api_rows(current)
                cleaned_rows, error_rows, unmatched = validate_and_clean(rows)
                output_path = write_clean_output(cleaned_rows, error_rows, "Configured API")
                show_results(cleaned_rows, error_rows, unmatched, output_path)
            except Exception as exc:
                st.error(f"API processing failed: {exc}")


def main():
    st.set_page_config(page_title="BDO Reimbursement Cleanup", layout="wide")
    st.title("BDO Reimbursement Cleanup Tool")
    st.write("Validate, clean, and export reimbursement data into the required BDO upload template.")

    if not os.path.exists(TEMPLATE_PATH):
        st.error("Missing reference/bdo_template.xlsm in the GitHub repository.")
        return

    tab1, tab2 = st.tabs(["Upload File", "Smartsheet/API"])
    with tab1:
        upload_tab()
    with tab2:
        api_tab()


main()
