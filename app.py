import os
import re
import csv
import uuid
import unicodedata
import json
import base64
import urllib.request
import urllib.error
from decimal import Decimal, InvalidOperation
from datetime import datetime

from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "reference", "bdo_template.xlsm")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
CONFIG_PATH = os.path.join(BASE_DIR, "api_config.json")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("APP_SECRET_KEY", "dev-only-change-me")

DATA_START_ROW = 12
TEMPLATE_SHEET = "BOB Outward Payments"
BANK_CODES_SHEET = "Bank Codes"

COLUMNS = [
    "Amount", "Type of Beneficiary", "Beneficiary First Name", "Beneficiary Middle Name",
    "Beneficiary Last Name or Corporate Beneficiary Name", "Beneficiary Address 1",
    "Beneficiary Address 2", "Beneficiary Address 3", "Beneficiary Account Number",
    "Beneficiary Bank", "Beneficiary Bank Address", "Beneficiary Information",
    "Bank to Bank Information", "Purpose Code", "Nature of Transfer", "Swift Code",
    "Country of Destination", "Importers Code", "Routing Number", "RTGS Purpose Code",
]

MANDATORY_FIELDS = [
    "Amount", "Type of Beneficiary", "Beneficiary Last Name or Corporate Beneficiary Name",
    "Beneficiary Address 1", "Beneficiary Address 2", "Beneficiary Address 3",
    "Beneficiary Account Number", "Beneficiary Bank", "Beneficiary Bank Address",
    "Purpose Code", "Country of Destination",
]

CLEAN_TEXT_FIELDS = [
    "Beneficiary First Name", "Beneficiary Middle Name",
    "Beneficiary Last Name or Corporate Beneficiary Name", "Beneficiary Address 1",
    "Beneficiary Address 2", "Beneficiary Address 3", "Beneficiary Bank Address",
    "Beneficiary Information", "Bank to Bank Information", "Nature of Transfer",
    "Country of Destination", "Swift Code", "Importers Code", "Routing Number",
]
ADDRESS_FIELDS = ["Beneficiary Address 1", "Beneficiary Address 2", "Beneficiary Address 3"]

# Handles your monitoring CSV headers and common BDO header variations.
HEADER_ALIASES = {
    "account number": "Beneficiary Account Number",
    "beneficiary account number": "Beneficiary Account Number",
    "beneficiary last nameor corporate beneficiary nam": "Beneficiary Last Name or Corporate Beneficiary Name",
    "beneficiary last name or corporate beneficiary name": "Beneficiary Last Name or Corporate Beneficiary Name",
    "beneficiary bank": "Beneficiary Bank",
    "beneficiary bank address": "Beneficiary Bank Address",
    "bank to bank information": "Bank to Bank Information",
    "country of destination": "Country of Destination",
    "rtgs purpose code": "RTGS Purpose Code",
    "swift code": "Swift Code",
    "purpose code": "Purpose Code",
}

BANK_ALIASES = {
    "METROBANK": "METROPOLITAN BANK AND TRUST CO",
    "METROPOLITAN BANK": "METROPOLITAN BANK AND TRUST CO",
    "BPI": "BANK OF THE PHILIPPINE ISLANDS",
    "RCBC": "RIZAL COMMERCIAL BANKING CORP",
    "SECURITY BANK": "SECURITY BANK CORPORATION",
    "CHINABANK SAVINGS": "CHINA BANK SAVINGS",
    "CHINA BANK SAVINGS": "CHINA BANK SAVINGS",
    "CHINABANK": "CHINA BANKING CORPORATION",
    "BDO": "BDO UNIBANK, INC.",
    "BDO UNIBANK": "BDO UNIBANK, INC.",
    "GCASH": "GCASH",
    "METRO BANK": "METROPOLITAN BANK AND TRUST CO",
    "BANK OF THE PHILIPPINE ISLAND": "BANK OF THE PHILIPPINE ISLANDS",
    "BANK OF THE PHILIPPINE ISLAND BPI": "BANK OF THE PHILIPPINE ISLANDS",
    "BPI BANK": "BANK OF THE PHILIPPINE ISLANDS",
    "HELLOMONEY AUB": "ASIA UNITED BANK",
    "MAYA PHILIPPINES": "PAYMAYA PHILIPPINES INC",
    "PAYMAYA": "PAYMAYA PHILIPPINES INC",
    "PS BANK": "PHILIPPINE SAVINGS BANK",
    "RCBC DISKARTECH": "RIZAL COMMERCIAL BANKING CORP",
    "RCBC SAVINGS BANK": "RIZAL COMMERCIAL BANKING CORP",
    "UNION BANK": "UNION BANK OF THE PHILIPPINES",
    "UNIONBANK": "UNION BANK OF THE PHILIPPINES",
    "UNIONBANK OF THE PHILIPPINES": "UNION BANK OF THE PHILIPPINES",
}

IGNORED_COLUMNS = {
    "date endorsed",
    "employer name",
    "status",
    "tracking no",
    "tracking no.",
    "upload date",
    "uploaded",
}



def load_api_config():
    if not os.path.exists(CONFIG_PATH):
        return {
            "api_url": "",
            "method": "GET",
            "auth_type": "none",
            "bearer_token": "",
            "basic_username": "",
            "basic_password": "",
            "custom_headers": "",
            "json_data_path": "",
        }
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception:
        config = {}
    defaults = load_api_config.__defaults__[0] if load_api_config.__defaults__ else None
    base = {
        "api_url": "",
        "method": "GET",
        "auth_type": "none",
        "bearer_token": "",
        "basic_username": "",
        "basic_password": "",
        "custom_headers": "",
        "json_data_path": "",
    }
    base.update(config)
    return base


def save_api_config(config):
    safe = {k: config.get(k, "") for k in [
        "api_url", "method", "auth_type", "bearer_token", "basic_username",
        "basic_password", "custom_headers", "json_data_path"
    ]}
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(safe, f, indent=2)


def parse_custom_headers(text):
    headers = {}
    for line in (text or "").splitlines():
        if not line.strip() or ":" not in line:
            continue
        key, value = line.split(":", 1)
        headers[key.strip()] = value.strip()
    return headers


def extract_json_path(data, path):
    current = data
    path = (path or "").strip()
    if not path:
        return current
    for part in path.split("."):
        part = part.strip()
        if not part:
            continue
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, list) and part.isdigit():
            current = current[int(part)]
        else:
            raise ValueError(f"JSON path not found at '{part}'")
    return current


def fetch_api_rows(config):
    api_url = (config.get("api_url") or "").strip()
    if not api_url:
        raise ValueError("API URL is required.")

    method = (config.get("method") or "GET").upper()
    headers = {"Accept": "application/json"}
    headers.update(parse_custom_headers(config.get("custom_headers")))

    auth_type = config.get("auth_type", "none")
    if auth_type == "bearer" and config.get("bearer_token"):
        headers["Authorization"] = f"Bearer {config.get('bearer_token')}"
    elif auth_type == "basic" and config.get("basic_username"):
        raw = f"{config.get('basic_username')}:{config.get('basic_password', '')}".encode("utf-8")
        headers["Authorization"] = "Basic " + base64.b64encode(raw).decode("ascii")

    req = urllib.request.Request(api_url, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8-sig")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")[:500]
        raise ValueError(f"API returned HTTP {exc.code}: {detail}")

    try:
        data = json.loads(body)
    except json.JSONDecodeError:
        # Allow APIs that return CSV text instead of JSON.
        return list(csv.DictReader(body.splitlines()))

    # Optional: allow the config page to point to a nested JSON array,
    # for example: data.items or response.rows
    data = extract_json_path(data, config.get("json_data_path"))

    # Smartsheet "Get Sheet" response:
    # {
    #   "columns": [{"id": ..., "title": ...}],
    #   "rows": [{"cells": [{"columnId": ..., "value": ...}]}]
    # }
    # Convert it into ordinary dict rows:
    # {"Amount": "...", "Beneficiary Bank": "..."}
    if isinstance(data, dict) and "columns" in data and "rows" in data:
        columns = {col.get("id"): col.get("title") for col in data.get("columns", [])}
        records = []

        for row in data.get("rows", []):
            record = {}
            for cell in row.get("cells", []):
                column_name = columns.get(cell.get("columnId"))
                if column_name:
                    record[column_name] = cell.get("displayValue", cell.get("value", ""))
            records.append(record)

        return records

    # Some APIs return {"rows": [...]} or {"data": [...]}.
    if isinstance(data, dict):
        for key in ("rows", "data", "items", "records", "results"):
            if isinstance(data.get(key), list):
                data = data[key]
                break

    if not isinstance(data, list):
        raise ValueError("API response must be a list of rows, a Smartsheet sheet response, or use JSON data path to point to the row list.")

    if not all(isinstance(row, dict) for row in data):
        raise ValueError("API row data must be a list of objects/dictionaries.")

    return data

def fix_mojibake(value):
    if value is None:
        return ""
    text = str(value)
    # Repairs common UTF-8 text that was opened as Windows-1252, e.g. BIÃ‘AN -> BIÑAN.
    if "Ã" in text or "Â" in text:
        try:
            text = text.encode("latin1").decode("utf-8")
        except Exception:
            pass
    return text


def remove_accents(value):
    text = fix_mojibake(value)
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


def normalize_header(value):
    text = remove_accents(value).replace("\xa0", " ").replace("\n", " ")
    text = re.sub(r"\([^)]*\)", "", text)
    return re.sub(r"[^A-Za-z0-9]+", " ", text).strip().lower()


def is_blank(value):
    return value is None or str(value).strip() == ""


def clean_text(value, max_len=None):
    if is_blank(value):
        return ""
    text = remove_accents(value).upper().strip()
    text = re.sub(r"[^A-Z0-9 ]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:max_len].rstrip() if max_len else text


def clean_account(value):
    if is_blank(value):
        return ""
    text = str(value).strip().replace(",", "")
    try:
        if re.fullmatch(r"[+-]?\d+(\.\d+)?[eE][+-]?\d+", text):
            return str(Decimal(text).quantize(Decimal("1")))
        if re.fullmatch(r"\d+\.0+", text):
            return text.split(".")[0]
    except (InvalidOperation, ValueError):
        pass
    return re.sub(r"\D+", "", text)


def clean_amount(value):
    if is_blank(value):
        return ""
    try:
        return float(Decimal(str(value).replace(",", "").strip()))
    except Exception:
        return value


def canonical_lookup():
    lookup = {normalize_header(c): c for c in COLUMNS}
    lookup.update(HEADER_ALIASES)
    return lookup


def load_bank_names():
    wb = load_workbook(TEMPLATE_PATH, read_only=True, keep_vba=True, data_only=True)
    ws = wb[BANK_CODES_SHEET]
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and row[1]:
            names.add(clean_text(row[1]))
    return names


def canonical_bank(value):
    name = clean_text(value)
    return clean_text(BANK_ALIASES.get(name, name))


def read_csv_rows(path):
    encodings = ["utf-8-sig", "cp1252", "latin1"]
    last_error = None
    for enc in encodings:
        try:
            with open(path, newline="", encoding=enc) as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error


def read_excel_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    lookup = canonical_lookup()
    best_row, best_score = 1, -1
    for row_num in range(1, min(ws.max_row, 25) + 1):
        values = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]
        score = sum(1 for v in values if normalize_header(v) in lookup)
        if score > best_score:
            best_row, best_score = row_num, score
    headers = [ws.cell(best_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(best_row + 1, ws.max_row + 1):
        record, has_any = {}, False
        for c, header in enumerate(headers, start=1):
            if header is None:
                continue
            value = ws.cell(r, c).value
            if value not in (None, ""):
                has_any = True
            record[str(header)] = value
        if has_any:
            rows.append(record)
    return rows


def read_input_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return read_csv_rows(path)
    return read_excel_rows(path)


def map_row_keys(row):
    lookup = canonical_lookup()
    mapped = {c: "" for c in COLUMNS}
    unmatched = []

    for key, value in row.items():
        normalized = normalize_header(key)

        if normalized in lookup:
            mapped[lookup[normalized]] = value
        elif str(key).strip() and normalized not in IGNORED_COLUMNS:
            unmatched.append(str(key))

    return mapped, unmatched

def validate_and_clean(rows):
    bank_names = load_bank_names()
    cleaned_rows, error_rows, all_unmatched = [], [], set()
    for idx, row in enumerate(rows, start=2):
        out, unmatched = map_row_keys(row)
        all_unmatched.update(unmatched)
        errors = []

        out["Amount"] = clean_amount(out["Amount"])
        out["Beneficiary Account Number"] = clean_account(out["Beneficiary Account Number"])
        out["Beneficiary Bank"] = canonical_bank(out["Beneficiary Bank"])
        out["Type of Beneficiary"] = clean_text(out["Type of Beneficiary"])
        out["Purpose Code"] = clean_text(out["Purpose Code"])
        out["RTGS Purpose Code"] = clean_text(out["RTGS Purpose Code"])
        for field in CLEAN_TEXT_FIELDS:
            out[field] = clean_text(out[field], 35 if field in ADDRESS_FIELDS else None)

        for field in MANDATORY_FIELDS:
            if is_blank(out[field]):
                errors.append(f"Missing mandatory field: {field}")

        benef_type = out["Type of Beneficiary"]
        if benef_type.startswith("IND") and is_blank(out["Beneficiary First Name"]):
            errors.append("Missing Beneficiary First Name for Individual beneficiary")
        if benef_type.startswith("CORP"):
            out["Beneficiary First Name"] = ""
            out["Beneficiary Middle Name"] = ""

        if not is_blank(out["Amount"]):
            try:
                if float(out["Amount"]) <= 0:
                    errors.append("Amount must be greater than zero")
            except Exception:
                errors.append("Amount must be numeric")

        if not is_blank(out["Beneficiary Bank"]) and out["Beneficiary Bank"] not in bank_names:
            errors.append("Beneficiary Bank not found in Bank Codes sheet")

        cleaned_rows.append(out)
        if errors:
            error_rows.append({"Source Row": idx, "Errors": "; ".join(errors), **out})
    return cleaned_rows, error_rows, sorted(all_unmatched)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)
    ws.freeze_panes = "A2"


def write_clean_output(cleaned_rows, error_rows, original_filename):
    wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb[TEMPLATE_SHEET]
    clear_until = max(ws.max_row, DATA_START_ROW + len(cleaned_rows) + 10)
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=clear_until, min_col=1, max_col=20):
        for cell in row:
            cell.value = None
    for r_idx, item in enumerate(cleaned_rows, start=DATA_START_ROW):
        for c_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(r_idx, c_idx).value = item.get(col, "")

    for sheet_name in ["Validation Errors", "Summary"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    err_ws = wb.create_sheet("Validation Errors")
    err_headers = ["Source Row", "Errors"] + COLUMNS
    err_ws.append(err_headers)
    for err in error_rows:
        err_ws.append([err.get(h, "") for h in err_headers])
    style_sheet(err_ws)

    sum_ws = wb.create_sheet("Summary", 0)
    sum_ws.append(["Metric", "Value"])
    sum_ws.append(["Source File", original_filename])
    sum_ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    sum_ws.append(["Total Rows Processed", len(cleaned_rows)])
    sum_ws.append(["Rows With Errors", len(error_rows)])
    sum_ws.append(["Ready Rows", len(cleaned_rows) - len(error_rows)])
    style_sheet(sum_ws)

    token = uuid.uuid4().hex[:8]
    output_path = os.path.join(OUTPUT_DIR, f"bdo_cleaned_upload_{token}.xlsm")
    wb.save(output_path)
    return output_path


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", api_config=load_api_config())


@app.route("/validate", methods=["POST"])
def validate():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Please upload a CSV, XLSX, or XLSM file.")
        return redirect(url_for("index"))
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".csv", ".xlsx", ".xlsm"}:
        flash("Unsupported file type. Please upload CSV, XLSX, or XLSM.")
        return redirect(url_for("index"))
    saved_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4().hex}_{filename}")
    file.save(saved_path)
    try:
        rows = read_input_file(saved_path)
        cleaned_rows, error_rows, unmatched = validate_and_clean(rows)
        output_path = write_clean_output(cleaned_rows, error_rows, filename)
    except Exception as exc:
        flash(f"Processing failed: {exc}")
        return redirect(url_for("index"))
    return render_template(
        "result.html",
        total=len(cleaned_rows),
        errors=len(error_rows),
        ready=len(cleaned_rows) - len(error_rows),
        unmatched=unmatched,
        download_name=os.path.basename(output_path),
        error_preview=error_rows[:20],
    )



@app.route("/api-config", methods=["POST"])
def api_config():
    config = {
        "api_url": request.form.get("api_url", ""),
        "method": request.form.get("method", "GET"),
        "auth_type": request.form.get("auth_type", "none"),
        "bearer_token": request.form.get("bearer_token", ""),
        "basic_username": request.form.get("basic_username", ""),
        "basic_password": request.form.get("basic_password", ""),
        "custom_headers": request.form.get("custom_headers", ""),
        "json_data_path": request.form.get("json_data_path", ""),
    }
    save_api_config(config)
    flash("API configuration saved.")
    return redirect(url_for("index"))


@app.route("/validate-api", methods=["POST"])
def validate_api():
    config = load_api_config()
    try:
        rows = fetch_api_rows(config)
        cleaned_rows, error_rows, unmatched = validate_and_clean(rows)
        source_name = config.get("api_url", "Configured API")
        output_path = write_clean_output(cleaned_rows, error_rows, "API: " + source_name)
    except Exception as exc:
        flash(f"API processing failed: {exc}")
        return redirect(url_for("index"))
    return render_template(
        "result.html",
        total=len(cleaned_rows),
        errors=len(error_rows),
        ready=len(cleaned_rows) - len(error_rows),
        unmatched=unmatched,
        download_name=os.path.basename(output_path),
        error_preview=error_rows[:20],
    )

@app.route("/download/<filename>")
def download(filename):
    path = os.path.join(OUTPUT_DIR, secure_filename(filename))
    return send_file(path, as_attachment=True, download_name=filename)


if __name__ == "__main__":
    app.run(debug=True)
