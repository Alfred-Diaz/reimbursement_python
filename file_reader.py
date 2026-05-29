import csv
import os

from openpyxl import load_workbook

from utils import canonical_lookup, normalize_header


def read_csv_rows(path):
    encodings = ["utf-8-sig", "cp1252", "latin1"]
    last_error = None

    for encoding in encodings:
        try:
            with open(path, newline="", encoding=encoding) as file:
                return list(csv.DictReader(file))
        except UnicodeDecodeError as exc:
            last_error = exc

    raise last_error


def read_excel_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    lookup = canonical_lookup()
    best_row, best_score = 1, -1

    for row_num in range(1, min(worksheet.max_row, 25) + 1):
        values = [worksheet.cell(row_num, col).value for col in range(1, worksheet.max_column + 1)]
        score = sum(1 for value in values if normalize_header(value) in lookup)
        if score > best_score:
            best_row, best_score = row_num, score

    headers = [worksheet.cell(best_row, col).value for col in range(1, worksheet.max_column + 1)]
    rows = []

    for row_num in range(best_row + 1, worksheet.max_row + 1):
        record, has_any = {}, False
        for col, header in enumerate(headers, start=1):
            if header is None:
                continue
            value = worksheet.cell(row_num, col).value
            if value not in (None, ""):
                has_any = True
            record[str(header)] = value
        if has_any:
            rows.append(record)

    return rows


def read_input_file(path):
    extension = os.path.splitext(path)[1].lower()
    if extension == ".csv":
        return read_csv_rows(path)
    return read_excel_rows(path)
