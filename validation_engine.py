from openpyxl import load_workbook

from constants import (
    ADDRESS_FIELDS,
    BANK_CODES_SHEET,
    CLEAN_TEXT_FIELDS,
    MANDATORY_FIELDS,
    TEMPLATE_PATH,
)
from utils import (
    canonical_bank,
    clean_account,
    clean_amount,
    clean_text,
    is_blank,
    map_row_keys,
)

PRESERVE_TEXT_FIELDS = {
    "Beneficiary Information",
    "Bank to Bank Information",
}


def preserve_reference_value(value):
    if value is None:
        return ""
    return str(value).strip()


def load_bank_names():
    workbook = load_workbook(
        TEMPLATE_PATH,
        read_only=True,
        keep_vba=True,
        data_only=True,
    )
    worksheet = workbook[BANK_CODES_SHEET]

    names = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and row[1]:
            names.add(clean_text(row[1]))

    return names


def validate_and_clean(rows):
    bank_names = load_bank_names()

    cleaned_rows = []
    error_rows = []
    unmatched_columns = set()

    for index, row in enumerate(rows, start=2):
        output, unmatched = map_row_keys(row)
        unmatched_columns.update(unmatched)

        errors = []

        output["Amount"] = clean_amount(output["Amount"])
        output["Beneficiary Account Number"] = clean_account(
            output["Beneficiary Account Number"]
        )
        output["Beneficiary Bank"] = canonical_bank(
            output["Beneficiary Bank"]
        )
        output["Type of Beneficiary"] = clean_text(
            output["Type of Beneficiary"]
        )
        output["Purpose Code"] = clean_text(output["Purpose Code"])
        output["RTGS Purpose Code"] = clean_text(
            output["RTGS Purpose Code"]
        )

        for field in CLEAN_TEXT_FIELDS:
            if field in PRESERVE_TEXT_FIELDS:
                output[field] = preserve_reference_value(output[field])
                continue
            limit = 35 if field in ADDRESS_FIELDS else None
            output[field] = clean_text(output[field], limit)

        for field in MANDATORY_FIELDS:
            if is_blank(output[field]):
                errors.append(f"Missing mandatory field: {field}")

        beneficiary_type = output["Type of Beneficiary"]

        if beneficiary_type.startswith("IND") and is_blank(
            output["Beneficiary First Name"]
        ):
            errors.append(
                "Missing Beneficiary First Name for Individual beneficiary"
            )

        if beneficiary_type.startswith("CORP"):
            output["Beneficiary First Name"] = ""
            output["Beneficiary Middle Name"] = ""

        if not is_blank(output["Amount"]):
            try:
                if float(output["Amount"]) <= 0:
                    errors.append("Amount must be greater than zero")
            except Exception:
                errors.append("Amount must be numeric")

        if (
            not is_blank(output["Beneficiary Bank"])
            and output["Beneficiary Bank"] not in bank_names
        ):
            errors.append(
                "Beneficiary Bank not found in Bank Codes sheet"
            )

        cleaned_rows.append(output)

        if errors:
            error_rows.append(
                {
                    "Source Row": index,
                    "Errors": "; ".join(errors),
                    **output,
                }
            )

    return cleaned_rows, error_rows, sorted(unmatched_columns)
