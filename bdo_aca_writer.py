import os
import uuid

from openpyxl import load_workbook

from constants import (
    BDO_ACA_COLUMNS,
    BDO_ACA_DATA_START_ROW,
    BDO_ACA_SHEET,
    BDO_ACA_TEMPLATE_PATH,
    OUTPUT_DIR,
)
from utils import clean_account, clean_amount, clean_text, is_blank


def build_particulars(row):
    first_name = clean_text(row.get("Beneficiary First Name", ""))
    middle_name = clean_text(row.get("Beneficiary Middle Name", ""))
    last_or_company = clean_text(
        row.get("Beneficiary Last Name or Corporate Beneficiary Name", "")
    )

    parts = [first_name, middle_name, last_or_company]
    name = " ".join(part for part in parts if part).strip()

    if name:
        return name

    return clean_text(row.get("Beneficiary Information", ""))


def map_to_bdo_aca_row(row):
    return {
        "Account Number": clean_account(row.get("Beneficiary Account Number", "")),
        "Reference Number": clean_text(row.get("Beneficiary Information", "")),
        "Client Transaction Number": clean_text(row.get("Bank to Bank Information", "")),
        "Amount": clean_amount(row.get("Amount", "")),
        "Particulars": build_particulars(row),
        "Email": clean_text(row.get("Email", "")),
        "Remarks": "",
    }


def validate_bdo_aca_rows(rows):
    errors = []

    for index, row in enumerate(rows, start=2):
        account_number = clean_account(row.get("Beneficiary Account Number", ""))
        amount = clean_amount(row.get("Amount", ""))

        row_errors = []
        if is_blank(account_number):
            row_errors.append("Missing BDO account number")
        if is_blank(amount):
            row_errors.append("Missing amount")
        else:
            try:
                if float(amount) <= 0:
                    row_errors.append("Amount must be greater than zero")
            except Exception:
                row_errors.append("Amount must be numeric")

        if row_errors:
            errors.append(
                {
                    "Source Row": index,
                    "Errors": "; ".join(row_errors),
                    **row,
                }
            )

    return errors


def clear_bdo_aca_template(worksheet, row_count):
    clear_until = max(
        worksheet.max_row,
        BDO_ACA_DATA_START_ROW + row_count + 10,
    )

    for row in worksheet.iter_rows(
        min_row=BDO_ACA_DATA_START_ROW,
        max_row=clear_until,
        min_col=1,
        max_col=len(BDO_ACA_COLUMNS),
    ):
        for cell in row:
            cell.value = None


def write_bdo_aca_output(rows, original_filename="BDO ACA"):
    if not os.path.exists(BDO_ACA_TEMPLATE_PATH):
        raise FileNotFoundError(
            "Missing reference/bdo_aca_template.xlsm in the GitHub repository."
        )

    workbook = load_workbook(BDO_ACA_TEMPLATE_PATH, keep_vba=True)
    worksheet = workbook[BDO_ACA_SHEET]

    mapped_rows = [map_to_bdo_aca_row(row) for row in rows]
    validation_errors = validate_bdo_aca_rows(rows)

    clear_bdo_aca_template(worksheet, len(mapped_rows))

    for row_index, item in enumerate(mapped_rows, start=BDO_ACA_DATA_START_ROW):
        for column_index, column_name in enumerate(BDO_ACA_COLUMNS, start=1):
            worksheet.cell(row_index, column_index).value = item.get(column_name, "")

    token = uuid.uuid4().hex[:8]
    output_path = os.path.join(
        OUTPUT_DIR,
        f"bdo_aca_upload_{token}.xlsm",
    )
    workbook.save(output_path)

    return output_path, validation_errors
