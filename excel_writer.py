import os
import uuid
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from constants import (
    COLUMNS,
    DATA_START_ROW,
    OUTPUT_DIR,
    TEMPLATE_PATH,
    TEMPLATE_SHEET,
)


def style_sheet(worksheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max(max_len + 2, 12),
            45,
        )

    worksheet.freeze_panes = "A2"


def clear_template_data(worksheet, row_count):
    clear_until = max(
        worksheet.max_row,
        DATA_START_ROW + row_count + 10,
    )

    for row in worksheet.iter_rows(
        min_row=DATA_START_ROW,
        max_row=clear_until,
        min_col=1,
        max_col=20,
    ):
        for cell in row:
            cell.value = None


def write_main_template_sheet(worksheet, cleaned_rows):
    clear_template_data(worksheet, len(cleaned_rows))

    for row_index, item in enumerate(cleaned_rows, start=DATA_START_ROW):
        for column_index, column_name in enumerate(COLUMNS, start=1):
            worksheet.cell(row_index, column_index).value = item.get(
                column_name,
                "",
            )


def write_validation_errors_sheet(workbook, error_rows):
    sheet_name = "Validation Errors"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    worksheet = workbook.create_sheet(sheet_name)
    headers = ["Source Row", "Errors"] + COLUMNS
    worksheet.append(headers)

    for error in error_rows:
        worksheet.append([error.get(header, "") for header in headers])

    style_sheet(worksheet)


def write_summary_sheet(workbook, original_filename, cleaned_rows, error_rows):
    sheet_name = "Summary"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    worksheet = workbook.create_sheet(sheet_name, 0)
    worksheet.append(["Metric", "Value"])
    worksheet.append(["Source File", original_filename])
    worksheet.append([
        "Generated At",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    worksheet.append(["Total Rows Processed", len(cleaned_rows)])
    worksheet.append(["Rows With Errors", len(error_rows)])
    worksheet.append([
        "Ready Rows",
        len(cleaned_rows) - len(error_rows),
    ])

    style_sheet(worksheet)


def write_clean_output(cleaned_rows, error_rows, original_filename):
    workbook = load_workbook(TEMPLATE_PATH, keep_vba=True)
    worksheet = workbook[TEMPLATE_SHEET]

    write_main_template_sheet(worksheet, cleaned_rows)
    write_validation_errors_sheet(workbook, error_rows)
    write_summary_sheet(
        workbook,
        original_filename,
        cleaned_rows,
        error_rows,
    )

    token = uuid.uuid4().hex[:8]
    output_path = os.path.join(
        OUTPUT_DIR,
        f"bdo_cleaned_upload_{token}.xlsm",
    )
    workbook.save(output_path)
    return output_path
